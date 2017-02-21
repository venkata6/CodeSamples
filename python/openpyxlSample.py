from openpyxl import load_workbook
import base64
import sys

def main(argv):
    print argv
    wb = load_workbook(argv[1])
    ws = wb.active
    rowCnt=0
    #print wb2.get_sheet_names()
    for row in ws.iter_rows():
        if row[1].value:  # if the column has value for this row
            if rowCnt!=0:   #dont change the heading row 
                tmp=base64.b64decode(row[1].value)
                str = unicode(tmp, errors='replace')
                print str
                row[1].value=str
            rowCnt+=1
            
    wb.save('NetflixReportNew.xlsx')

if __name__ == "__main__":
    if sys.argv[1:]:
        main(sys.argv)
    else:
        print "Usage: python decodeCIB <netflixExcelFileName>"
